import time
import threading
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook

# ---------------- FUNCIONES SELENIUM ----------------
def encontrar_regiones(driver):
    xpaths = [
        "//h1/following-sibling::ul",
        "//h1/following-sibling::p/following-sibling::ul",
        "//h2/following-sibling::ul",
        "//h2/following-sibling::p/following-sibling::ul",
        "//div//ul",
    ]

    for xp in xpaths:
        try:
            ul = driver.find_element(By.XPATH, xp)
            enlaces = ul.find_elements(By.TAG_NAME, "a")
            if len(enlaces) > 0:
                return enlaces
        except:
            pass

    return None


def encontrar_ul_flexible(driver):
    posibles_xpaths = [
        "//h2/following-sibling::p/following-sibling::ul",
        "//h2/following-sibling::ul",
        "//h1/following-sibling::p/following-sibling::ul",
        "//h1/following-sibling::ul"
    ]
    for xp in posibles_xpaths:
        try:
            return driver.find_element(By.XPATH, xp)
        except:
            continue
    return None

# ---------------- FUNCIÓN PRINCIPAL ----------------
def iniciar_scraping(rubro_input, output_box, boton):
    threading.Thread(target=lambda: ejecutar_scraping(rubro_input, output_box, boton)).start()

def ejecutar_scraping(rubro_input, output_box, boton):
    BASE_URL = "https://www.universidadperu.com/empresas/categorias.php"

    boton.config(state="disabled")
    output_box.insert(tk.END, f"Iniciando búsqueda del rubro '{rubro_input}'...\n")
    output_box.see(tk.END)
    output_box.update()

    filename = f"empresas_{rubro_input.replace(' ', '_')}.xlsx"

    try:
        wb = load_workbook(filename)
    except:
        wb = Workbook()
        wb.active.append(["Región", "Provincia", "Distrito", "Razón Social", "URL_Empresa"])
        wb.save(filename)

    wb = load_workbook(filename)
    ws = wb.active

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    wait = WebDriverWait(driver, 10)

    driver.get(BASE_URL)
    time.sleep(1)

    categoria_url = None
    for a in driver.find_elements(By.CSS_SELECTOR, "a"):
        texto = a.text.strip().lower()
        if rubro_input in texto:
            categoria_url = a.get_attribute("href")
            output_box.insert(tk.END, f"✅ Categoría encontrada: {texto}\n")
            output_box.see(tk.END)
            output_box.update()
            break

    if not categoria_url:
        output_box.insert(tk.END, f"❌ No se encontró la categoría '{rubro_input}'.\n")
        output_box.see(tk.END)
        driver.quit()
        boton.config(state="normal")
        return

    driver.get(categoria_url)
    time.sleep(1)

    regiones = encontrar_regiones(driver)
    if not regiones:
        output_box.insert(tk.END, "❌ No se pudieron obtener las regiones.\n")
        output_box.see(tk.END)
        driver.quit()
        boton.config(state="normal")
        return

    regiones_data = [(r.text.strip(), r.get_attribute("href")) for r in regiones if r.text.strip()]

    output_box.insert(tk.END, f"📍 Se encontraron {len(regiones_data)} regiones.\n\n")
    output_box.see(tk.END)

    # ---------------- INICIO DEL SCRAPING ----------------
    for nombre_region, url_region in regiones_data:
        output_box.insert(tk.END, f" Región: {nombre_region}\n")
        output_box.see(tk.END)
        output_box.update()

        driver.get(url_region)
        time.sleep(1)

        ul_prov = encontrar_ul_flexible(driver)
        if not ul_prov:
            output_box.insert(tk.END, "   ⚠️ No se encontró bloque de provincias.\n")
            output_box.see(tk.END)
            continue

        provincias = ul_prov.find_elements(By.TAG_NAME, "a")
        provincias_data = [(p.text.strip(), p.get_attribute("href")) for p in provincias if p.text.strip()]

        for nombre_prov, url_prov in provincias_data:
            output_box.insert(tk.END, f"    Provincia: {nombre_prov}\n")
            output_box.see(tk.END)
            output_box.update()

            driver.get(url_prov)
            time.sleep(1)

            ul_dist = encontrar_ul_flexible(driver)
            if not ul_dist:
                output_box.insert(tk.END, "      ⚠️ No se encontró bloque de distritos.\n")
                output_box.see(tk.END)
                continue

            distritos = ul_dist.find_elements(By.TAG_NAME, "a")
            distritos_data = [(d.text.strip(), d.get_attribute("href")) for d in distritos if d.text.strip()]

            for nombre_dist, url_dist in distritos_data:
                output_box.insert(tk.END, f"       Distrito: {nombre_dist}\n")
                output_box.see(tk.END)
                output_box.update()

                driver.get(url_dist)
                time.sleep(1)

                ul_emp = encontrar_ul_flexible(driver)
                if not ul_emp:
                    output_box.insert(tk.END, "         ⚠️ No se encontró bloque de empresas.\n")
                    output_box.see(tk.END)
                    continue

                empresas = ul_emp.find_elements(By.TAG_NAME, "a")
                for e in empresas:
                    nombre_empresa = e.text.strip()
                    url_empresa = e.get_attribute("href")

                    if nombre_empresa and url_empresa:
                        ws.append([nombre_region, nombre_prov, nombre_dist, nombre_empresa, url_empresa])
                        wb.save(filename)

                        output_box.insert(tk.END, f"          {nombre_empresa}\n")
                        output_box.see(tk.END)
                        output_box.update()

    driver.quit()

    messagebox.showinfo("Éxito", f"Scraping terminado.\nArchivo guardado: {filename}")
    boton.config(state="normal")


# ---------------- INTERFAZ TKINTER ----------------
def crear_interfaz():
    root = tk.Tk()
    root.title("Scraper UniversidadPeru")
    root.geometry("700x500")

    frame = ttk.Frame(root, padding=20)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="🔎 Ingresa el rubro de empresas:").pack(pady=5)
    rubro_entry = ttk.Entry(frame, width=40)
    rubro_entry.pack(pady=5)

    output_box = tk.Text(frame, height=20, wrap="word")
    output_box.pack(fill="both", expand=True, pady=10)

    boton = ttk.Button(frame, text="Iniciar scraping",
                       command=lambda: iniciar_scraping(
                           rubro_entry.get().strip().lower(), output_box, boton
                       ))
    boton.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    crear_interfaz()
